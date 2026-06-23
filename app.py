"""
Resource Capacity Planning
==========================================================================
A board-ready Streamlit application for forecasting team working-day
capacity across the second half of the year (July–December), mirroring the
architecture of the "Replan H2" tab in Teams.xlsx.

Data model (mirrors Replan H2 columns A–N)
    A  Name           -> Author
    B  Team           -> Team       (SWE, SWA, Cloud, QA, Load, Fix, ...)
    C  TL             -> Team Lead
    D  Code           -> Work stream / code base
    E  Capacity       -> "%" column. 1 = 100%, 0.5 = 50%, 0 = NOT contributing.
    F  Contract       -> Contract type (FT, "4 days/ week", Contractor, ...)
    G  StartDate      -> Start date  (drives the new-starter ramp-up)
    H  EndDate        -> End date    (blank / "N/A" = open ended)
    I..N  July..December -> computed available working days per month

Persistence
    Primary backend is a Google Sheet accessed through
    st.connection("gsheets", type=GSheetsConnection). If the connection is
    not yet configured the app transparently falls back to an in-session
    store so it is always runnable and demoable. Connection status is shown
    in the sidebar.

Capacity logic
    available_days(month) =
        ( MaxWorkingDays
          - SicknessAllowance
          - HolidayAllowance
          - AdminSCDays )            # global, set via sidebar sliders
        * Capacity                    # the "%" column (0 => 0 days)
        * RampFactor(StartDate,month) # 33% / 66% / 100% for months 1/2/3
        - LoggedAbsenceDays(month)    # individual holidays from the tracker
    ...floored at 0 and rounded to 2 dp. A result of 0 means the person is
    not contributing that month, exactly as 0 is used on the Replan H2 tab.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import os
from typing import Optional

import pandas as pd
import streamlit as st

# Google Sheets connection is optional at import time so the app still runs
# (with the in-session fallback) on machines where it is not installed.
try:
    from streamlit_gsheets import GSheetsConnection
    _GSHEETS_IMPORTED = True
except Exception:  # pragma: no cover - import guard
    GSheetsConnection = None  # type: ignore
    _GSHEETS_IMPORTED = False


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

APP_TITLE = "Resource Capacity Planning · H2 Replan"

# Forecast months. (Excel column letter, label, calendar month number).
# Columns I..N on the Replan H2 tab == July..December.
MONTHS = [
    ("I", "July", 7),
    ("J", "August", 8),
    ("K", "September", 9),
    ("L", "October", 10),
    ("M", "November", 11),
    ("N", "December", 12),
]
MONTH_LABELS = [m[1] for m in MONTHS]
MONTH_NUMS = {m[1]: m[2] for m in MONTHS}

# Defaults for the global sidebar controls.
DEFAULT_MAX_DAYS = 15.5      # baseline maximum working days per person / month
DEFAULT_SICKNESS = 0.5       # sickness allowance deducted per month
DEFAULT_HOLIDAY = 2.0        # standard holiday baseline deduction per month
DEFAULT_ADMIN_SC = 1.0       # admin / Strategy & Culture days per month

# Ramp-up schedule: month 1 / 2 / 3 from the start date.
RAMP_SCHEDULE = {1: 0.33, 2: 0.66, 3: 1.00}

TEAM_COLUMNS = ["Name", "Team", "TL", "Code", "Capacity",
                "Contract", "StartDate", "EndDate"]
HOLIDAY_COLUMNS = ["Name", "Month", "Days", "Note"]

TEAM_WS = "Team"          # worksheet holding employee profiles
HOLIDAY_WS = "Holidays"   # worksheet holding the absence log

KNOWN_TEAMS = ["SWE", "SWA", "Cloud", "QA", "Load", "Fix"]
KNOWN_CONTRACTS = ["FT", "Contractor", "4 days/ week", "4.5 days/week", "Part-Time"]


# ---------------------------------------------------------------------------
# Persistence layer (Google Sheets primary, in-session fallback)
# ---------------------------------------------------------------------------

def _empty_team_df() -> pd.DataFrame:
    return pd.DataFrame(columns=TEAM_COLUMNS)


def _empty_holiday_df() -> pd.DataFrame:
    return pd.DataFrame(columns=HOLIDAY_COLUMNS)


@st.cache_resource(show_spinner=False)
def _get_connection():
    """Return a cached GSheets connection, or None if unavailable."""
    if not _GSHEETS_IMPORTED:
        return None
    try:
        return st.connection("gsheets", type=GSheetsConnection)
    except Exception:
        return None


def _coerce_team_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise a roster dataframe to the canonical schema/types."""
    df = df.copy()
    for col in TEAM_COLUMNS:
        if col not in df.columns:
            df[col] = "" if col != "Capacity" else 1.0
    df = df[TEAM_COLUMNS]
    # Drop fully blank rows (the data_editor leaves trailing empties).
    df = df[df["Name"].astype(str).str.strip() != ""]
    df["Capacity"] = pd.to_numeric(df["Capacity"], errors="coerce").fillna(1.0)
    for c in ["Name", "Team", "TL", "Code", "Contract", "StartDate", "EndDate"]:
        df[c] = df[c].astype(str).replace({"nan": "", "NaT": "", "None": ""}).str.strip()
    return df.reset_index(drop=True)


def _coerce_holiday_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in HOLIDAY_COLUMNS:
        if col not in df.columns:
            df[col] = "" if col != "Days" else 0.0
    df = df[HOLIDAY_COLUMNS]
    df = df[df["Name"].astype(str).str.strip() != ""]
    df["Days"] = pd.to_numeric(df["Days"], errors="coerce").fillna(0.0)
    for c in ["Name", "Month", "Note"]:
        df[c] = df[c].astype(str).replace({"nan": "", "None": ""}).str.strip()
    return df.reset_index(drop=True)


def load_team_df() -> pd.DataFrame:
    """Read the roster from Google Sheets, falling back to the session store."""
    conn = _get_connection()
    if conn is not None:
        try:
            df = conn.read(worksheet=TEAM_WS, ttl=0)
            st.session_state["backend"] = "gsheets"
            return _coerce_team_df(df.dropna(how="all"))
        except Exception as exc:  # worksheet missing or auth issue
            st.session_state["backend"] = "session"
            st.session_state["backend_error"] = str(exc)
    if "team_df" not in st.session_state:
        st.session_state["team_df"] = _seed_team_df()
    st.session_state.setdefault("backend", "session")
    return _coerce_team_df(st.session_state["team_df"])


def load_holiday_df() -> pd.DataFrame:
    conn = _get_connection()
    if conn is not None:
        try:
            df = conn.read(worksheet=HOLIDAY_WS, ttl=0)
            return _coerce_holiday_df(df.dropna(how="all"))
        except Exception:
            pass
    if "holiday_df" not in st.session_state:
        st.session_state["holiday_df"] = _empty_holiday_df()
    return _coerce_holiday_df(st.session_state["holiday_df"])


def save_team_df(df: pd.DataFrame) -> None:
    """Persist the roster. Writes to Google Sheets when connected."""
    df = _coerce_team_df(df)
    conn = _get_connection()
    if conn is not None:
        try:
            conn.update(worksheet=TEAM_WS, data=df)
        except Exception as exc:
            st.warning(f"Could not write to Google Sheets ({exc}). "
                       "Changes are kept for this session only.")
    st.session_state["team_df"] = df


def save_holiday_df(df: pd.DataFrame) -> None:
    df = _coerce_holiday_df(df)
    conn = _get_connection()
    if conn is not None:
        try:
            conn.update(worksheet=HOLIDAY_WS, data=df)
        except Exception as exc:
            st.warning(f"Could not write holidays to Google Sheets ({exc}). "
                       "Changes are kept for this session only.")
    st.session_state["holiday_df"] = df


def _seed_team_df() -> pd.DataFrame:
    """Load the bundled roster (extracted from Teams.xlsx) if present."""
    seed_path = os.path.join(os.path.dirname(__file__), "seed_roster.json")
    try:
        with open(seed_path, "r", encoding="utf-8") as fh:
            return _coerce_team_df(pd.DataFrame(json.load(fh)))
    except Exception:
        return _empty_team_df()


# ---------------------------------------------------------------------------
# Capacity calculation engine
# ---------------------------------------------------------------------------

def _parse_date(value) -> Optional[dt.date]:
    """Parse a variety of date representations; '', 'N/A', NaT -> None."""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if s == "" or s.upper() == "N/A" or s.lower() in ("nat", "nan", "none"):
        return None
    try:
        return pd.to_datetime(s).date()
    except Exception:
        return None


def ramp_factor(start: Optional[dt.date], year: int, month: int) -> float:
    """Return the new-starter productivity factor for a calendar month.

    Months 1/2/3 counting from the start month ramp at 33% / 66% / 100%.
    Months before the start month return 0 (the person has not joined yet).
    A missing start date is treated as fully ramped (legacy team member).
    """
    if start is None:
        return 1.0
    months_since = (year * 12 + month) - (start.year * 12 + start.month) + 1
    if months_since <= 0:
        return 0.0
    return RAMP_SCHEDULE.get(months_since, 1.0)


def month_available_days(row: pd.Series, year: int, month: int,
                         max_days: float, sickness: float,
                         holiday_allow: float, admin_sc: float,
                         logged_absence: float) -> float:
    """Compute available working days for one person in one month."""
    capacity = float(row.get("Capacity", 0) or 0)
    # Capacity of 0 => not contributing this month (Replan H2 semantics).
    if capacity == 0:
        return 0.0

    start = _parse_date(row.get("StartDate"))
    end = _parse_date(row.get("EndDate"))

    rf = ramp_factor(start, year, month)
    if rf == 0.0:                       # not started yet
        return 0.0

    # Person has left before this month begins.
    if end is not None and end < dt.date(year, month, 1):
        return 0.0

    baseline = max_days - sickness - holiday_allow - admin_sc
    baseline = max(baseline, 0.0)
    available = baseline * capacity * rf
    available -= max(logged_absence, 0.0)
    return round(max(available, 0.0), 2)


def absence_lookup(holidays: pd.DataFrame) -> dict:
    """Aggregate the absence log into {(name, month): total_days}."""
    out: dict = {}
    if holidays.empty:
        return out
    grouped = holidays.groupby(["Name", "Month"])["Days"].sum()
    for (name, month), days in grouped.items():
        out[(str(name).strip(), str(month).strip())] = float(days)
    return out


def build_matrix(team: pd.DataFrame, holidays: pd.DataFrame, year: int,
                 max_days: float, sickness: float, holiday_allow: float,
                 admin_sc: float) -> pd.DataFrame:
    """Return the full forecast matrix: one row per person, columns per month."""
    absences = absence_lookup(holidays)
    records = []
    for _, row in team.iterrows():
        rec = {
            "Name": row["Name"], "Team": row["Team"], "TL": row["TL"],
            "Code": row["Code"], "Capacity": row["Capacity"],
            "Contract": row["Contract"], "StartDate": row["StartDate"],
            "EndDate": row["EndDate"],
        }
        for _, label, num in MONTHS:
            logged = absences.get((str(row["Name"]).strip(), label), 0.0)
            rec[label] = month_available_days(
                row, year, num, max_days, sickness,
                holiday_allow, admin_sc, logged)
        records.append(rec)
    cols = TEAM_COLUMNS + MONTH_LABELS
    return pd.DataFrame(records, columns=cols)


# ---------------------------------------------------------------------------
# Excel export — mirrors the Replan H2 architecture (team blocks + TOTAL/COUNT)
# ---------------------------------------------------------------------------

def build_excel(matrix: pd.DataFrame, year: int) -> bytes:
    """Build an .xlsx that mirrors the Replan H2 layout, using live formulas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Replan H2"

    header = ["Author", "Team", "TL", "Code", "%", "Contract type",
              "Start date", "End date"] + MONTH_LABELS
    ws.append(header)

    # ---- styling helpers -------------------------------------------------
    head_fill = PatternFill("solid", start_color="1F3864")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", start_color="D9E1F2")
    count_fill = PatternFill("solid", start_color="EDEDED")
    bold = Font(name="Arial", bold=True, size=10)
    base_font = Font(name="Arial", size=10)
    centre = Alignment(horizontal="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = (
            head_fill, head_font, centre, border)

    first_month_col = 9          # column I
    row_ptr = 2
    teams = list(dict.fromkeys(matrix["Team"].tolist()))  # preserve order

    for team in teams:
        block = matrix[matrix["Team"] == team]
        block_start = row_ptr
        for _, r in block.iterrows():
            ws.cell(row=row_ptr, column=1, value=r["Name"])
            ws.cell(row=row_ptr, column=2, value=r["Team"])
            ws.cell(row=row_ptr, column=3, value=r["TL"])
            ws.cell(row=row_ptr, column=4, value=r["Code"])
            ws.cell(row=row_ptr, column=5, value=float(r["Capacity"]))
            ws.cell(row=row_ptr, column=6, value=r["Contract"])
            ws.cell(row=row_ptr, column=7, value=r["StartDate"] or "")
            ws.cell(row=row_ptr, column=8, value=r["EndDate"] or "N/A")
            for i, label in enumerate(MONTH_LABELS):
                ws.cell(row=row_ptr, column=first_month_col + i,
                        value=float(r[label]))
            for c in range(1, len(header) + 1):
                cell = ws.cell(row=row_ptr, column=c)
                cell.font = base_font
                cell.border = border
                if c >= first_month_col or c == 5:
                    cell.alignment = centre
            row_ptr += 1
        block_end = row_ptr - 1

        # ----- TOTAL row (=SUM) -----
        ws.cell(row=row_ptr, column=1, value=f"{team} TOTAL").font = bold
        for i in range(len(MONTH_LABELS)):
            col = get_column_letter(first_month_col + i)
            cell = ws.cell(row=row_ptr, column=first_month_col + i)
            if block_end >= block_start:
                cell.value = f"=SUM({col}{block_start}:{col}{block_end})"
            else:
                cell.value = 0
            cell.font, cell.alignment = bold, centre
        for c in range(1, len(header) + 1):
            ws.cell(row=row_ptr, column=c).fill = total_fill
            ws.cell(row=row_ptr, column=c).border = border
        row_ptr += 1

        # ----- COUNT row (=COUNTIF "<>0") -- mirrors headcount logic -----
        ws.cell(row=row_ptr, column=1, value=f"{team} COUNT").font = bold
        for i in range(len(MONTH_LABELS)):
            col = get_column_letter(first_month_col + i)
            cell = ws.cell(row=row_ptr, column=first_month_col + i)
            if block_end >= block_start:
                cell.value = f'=COUNTIF({col}{block_start}:{col}{block_end},"<>0")'
            else:
                cell.value = 0
            cell.font, cell.alignment = bold, centre
        for c in range(1, len(header) + 1):
            ws.cell(row=row_ptr, column=c).fill = count_fill
            ws.cell(row=row_ptr, column=c).border = border
        row_ptr += 2  # blank spacer row between team blocks

    # ---- column widths & number formats ---------------------------------
    widths = {"A": 26, "B": 10, "C": 14, "D": 12, "E": 7, "F": 16,
              "G": 13, "H": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i in range(len(MONTH_LABELS)):
        ws.column_dimensions[get_column_letter(first_month_col + i)].width = 11
    ws.freeze_panes = "I2"

    for r in range(2, row_ptr):
        ws.cell(row=r, column=5).number_format = "0.0%"  # % column
        for i in range(len(MONTH_LABELS)):
            c = ws.cell(row=r, column=first_month_col + i)
            # 0 work-day months display as "-" exactly like a clean template.
            c.number_format = "0.0;[Red](0.0);-"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")

# Lightweight styling for a clean, board-ready feel.
st.markdown(
    """
    <style>
      .block-container {padding-top: 2.2rem;}
      [data-testid="stMetricValue"] {font-size: 1.6rem;}
      h1, h2, h3 {color: #1F3864;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📊 Resource Capacity Planning")
st.caption("H2 Replan · July–December working-day forecast — aligned to the "
           "Replan H2 model in Teams.xlsx")

# ---- Sidebar: global controls -------------------------------------------
with st.sidebar:
    st.header("⚙️ Global controls")

    forecast_year = st.number_input("Forecast year", min_value=2020,
                                    max_value=2100, value=2026, step=1)

    max_days = st.slider("Max working days / month", 0.0, 23.0,
                         DEFAULT_MAX_DAYS, 0.5,
                         help="Baseline maximum working days per person per month.")
    sickness = st.slider("Sickness allowance / month", 0.0, 10.0,
                         DEFAULT_SICKNESS, 0.5,
                         help="Deducted from each person's monthly days.")
    holiday_allow = st.slider("Holiday allowance / month", 0.0, 15.0,
                              DEFAULT_HOLIDAY, 0.5,
                              help="Standard baseline holiday deduction per month.")
    admin_sc = st.slider("Admin / S&C days / month", 0.0, 10.0,
                         DEFAULT_ADMIN_SC, 0.5,
                         help="Strategy & Culture / admin time deducted per month.")

    net_per_month = max(max_days - sickness - holiday_allow - admin_sc, 0.0)
    st.metric("Net productive days / month (100%)", f"{net_per_month:.1f}")

    st.divider()
    # Persistence status indicator.
    if not _GSHEETS_IMPORTED:
        st.warning("`st-gsheets-connection` not installed — running on the "
                   "in-session store. See the README to enable Google Sheets.")
    elif _get_connection() is None:
        st.warning("Google Sheets not configured — using the in-session store. "
                   "Add credentials to `.streamlit/secrets.toml`.")
    else:
        st.success("Connected to Google Sheets ✔")

# ---- Load data -----------------------------------------------------------
team_df = load_team_df()
holiday_df = load_holiday_df()

# Build the live forecast matrix used across all tabs.
matrix = build_matrix(team_df, holiday_df, int(forecast_year), max_days,
                      sickness, holiday_allow, admin_sc)

# ---- Headline metrics ----------------------------------------------------
total_days = float(matrix[MONTH_LABELS].sum().sum()) if not matrix.empty else 0.0
people_contributing = int(
    (matrix[MONTH_LABELS].sum(axis=1) > 0).sum()) if not matrix.empty else 0
team_count = matrix["Team"].nunique() if not matrix.empty else 0
avg_per_person = (total_days / people_contributing) if people_contributing else 0.0

m1, m2, m3, m4 = st.columns(4)
m1.metric("Total available days (H2)", f"{total_days:,.1f}")
m2.metric("People contributing", people_contributing)
m3.metric("Teams", team_count)
m4.metric("Avg days / person", f"{avg_per_person:,.1f}")

st.divider()

tab_team, tab_hol, tab_forecast = st.tabs(
    ["👥 Team & Capacity", "🏖️ Holidays & Absence", "📈 Forecast & Export"])

# =========================================================================
# TAB 1 — Team & Capacity
# =========================================================================
with tab_team:
    st.subheader("Add a team member")

    team_options = sorted(set(KNOWN_TEAMS) |
                          set([t for t in team_df["Team"].unique() if t]))

    with st.form("add_member", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            name = st.text_input("Name *")
            team_sel = st.selectbox("Team *", team_options + ["➕ New team…"])
            new_team = ""
            if team_sel == "➕ New team…":
                new_team = st.text_input("New team name")
        with c2:
            tl = st.text_input("Team Lead (TL)")
            code = st.text_input("Code / work stream")
            contract = st.selectbox("Contract type", KNOWN_CONTRACTS)
        with c3:
            # Capacity as a fully customisable numeric input (1.0 = 100%).
            cap_pct = st.number_input("Capacity (%)", min_value=0.0,
                                      max_value=200.0, value=100.0, step=10.0,
                                      help="100% = 1.0, 50% = 0.5, 0% = not "
                                           "contributing. Any value allowed.")
            start_date = st.date_input("Start date",
                                       value=dt.date(int(forecast_year), 7, 1))
            has_end = st.checkbox("Has an end date")
            end_date = st.date_input("End date", value=dt.date(
                int(forecast_year), 12, 31)) if has_end else None

        submitted = st.form_submit_button("➕ Add team member", type="primary")
        if submitted:
            final_team = new_team.strip() if team_sel == "➕ New team…" else team_sel
            if not name.strip():
                st.error("Name is required.")
            elif not final_team:
                st.error("Team is required.")
            else:
                new_row = {
                    "Name": name.strip(), "Team": final_team, "TL": tl.strip(),
                    "Code": code.strip(), "Capacity": round(cap_pct / 100.0, 4),
                    "Contract": contract,
                    "StartDate": start_date.strftime("%Y-%m-%d"),
                    "EndDate": end_date.strftime("%Y-%m-%d") if end_date else "",
                }
                updated = pd.concat(
                    [team_df, pd.DataFrame([new_row])], ignore_index=True)
                save_team_df(updated)              # instant write-back
                st.success(f"Added {name.strip()} to {final_team}.")
                st.rerun()

    st.subheader("Edit roster")
    st.caption("Capacity is stored as a fraction (1.0 = 100%). Edit inline, "
               "then click **Save roster**.")

    edited = st.data_editor(
        team_df,
        num_rows="dynamic",
        use_container_width=True,
        key="roster_editor",
        column_config={
            "Capacity": st.column_config.NumberColumn(
                "Capacity", help="1.0 = 100%, 0.5 = 50%, 0 = not contributing",
                min_value=0.0, max_value=2.0, step=0.05, format="%.2f"),
            "Team": st.column_config.TextColumn("Team"),
            "Contract": st.column_config.TextColumn("Contract"),
            "StartDate": st.column_config.TextColumn(
                "StartDate", help="YYYY-MM-DD"),
            "EndDate": st.column_config.TextColumn(
                "EndDate", help="YYYY-MM-DD or blank for open-ended"),
        },
    )
    cbtn1, cbtn2 = st.columns([1, 5])
    if cbtn1.button("💾 Save roster", type="primary"):
        save_team_df(edited)
        st.success("Roster saved.")
        st.rerun()

# =========================================================================
# TAB 2 — Holidays & Absence
# =========================================================================
with tab_hol:
    st.subheader("Log holiday / absence")
    st.caption("Logged days are deducted from that person's available working "
               "days for the chosen month.")

    if team_df.empty:
        st.info("Add team members first.")
    else:
        with st.form("add_holiday", clear_on_submit=True):
            h1, h2, h3, h4 = st.columns([2, 1.5, 1, 2])
            with h1:
                h_name = st.selectbox("Team member",
                                      sorted(team_df["Name"].unique()))
            with h2:
                h_month = st.selectbox("Month", MONTH_LABELS)
            with h3:
                h_days = st.number_input("Days", min_value=0.0, max_value=31.0,
                                         value=1.0, step=0.5)
            with h4:
                h_note = st.text_input("Note (optional)")
            h_submit = st.form_submit_button("➕ Log absence", type="primary")
            if h_submit:
                new_h = {"Name": h_name, "Month": h_month,
                         "Days": float(h_days), "Note": h_note.strip()}
                updated_h = pd.concat(
                    [holiday_df, pd.DataFrame([new_h])], ignore_index=True)
                save_holiday_df(updated_h)         # instant write-back
                st.success(f"Logged {h_days} day(s) for {h_name} in {h_month}.")
                st.rerun()

    st.subheader("Absence log")
    edited_h = st.data_editor(
        holiday_df,
        num_rows="dynamic",
        use_container_width=True,
        key="holiday_editor",
        column_config={
            "Month": st.column_config.SelectboxColumn(
                "Month", options=MONTH_LABELS),
            "Days": st.column_config.NumberColumn(
                "Days", min_value=0.0, max_value=31.0, step=0.5, format="%.1f"),
        },
    )
    if st.button("💾 Save absences", type="primary"):
        save_holiday_df(edited_h)
        st.success("Absences saved.")
        st.rerun()

    if not holiday_df.empty:
        st.subheader("Absence days by month")
        pivot = (holiday_df.groupby("Month")["Days"].sum()
                 .reindex(MONTH_LABELS).fillna(0.0))
        st.bar_chart(pivot)

# =========================================================================
# TAB 3 — Forecast & Export
# =========================================================================
with tab_forecast:
    st.subheader("Capacity matrix (available working days)")
    st.caption("Reflects capacity %, the 3-month new-starter ramp-up, global "
               "deductions and logged absences. 0 = not contributing — exactly "
               "as on the Replan H2 tab.")

    if matrix.empty:
        st.info("No team members yet — add some on the Team & Capacity tab.")
    else:
        show_cols = ["Name", "Team", "Capacity"] + MONTH_LABELS
        st.dataframe(
            matrix[show_cols],
            use_container_width=True, hide_index=True,
            column_config={lbl: st.column_config.NumberColumn(
                lbl, format="%.1f") for lbl in MONTH_LABELS},
        )

        st.subheader("Total available capacity by team / month")
        team_month = (matrix.groupby("Team")[MONTH_LABELS].sum()
                      .T.reindex(MONTH_LABELS))
        st.bar_chart(team_month)

        # Per-team totals table.
        st.subheader("Team totals (days)")
        totals = matrix.groupby("Team")[MONTH_LABELS].sum()
        totals["H2 Total"] = totals.sum(axis=1)
        totals.loc["All teams"] = totals.sum()
        st.dataframe(
            totals,
            use_container_width=True,
            column_config={c: st.column_config.NumberColumn(c, format="%.1f")
                           for c in totals.columns},
        )

        st.divider()
        st.subheader("Export")
        st.caption("Downloads an .xlsx mirroring the Replan H2 layout — team "
                   "blocks with =SUM TOTAL rows and =COUNTIF(\"<>0\") COUNT "
                   "rows, and 0-day months shown as \"-\".")
        xlsx_bytes = build_excel(matrix, int(forecast_year))
        st.download_button(
            "⬇️ Download Replan H2 matrix (.xlsx)",
            data=xlsx_bytes,
            file_name=f"Replan_H2_{int(forecast_year)}.xlsx",
            mime=("application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet"),
            type="primary",
        )
